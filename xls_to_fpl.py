from openpyxl import load_workbook
from xml.dom import minidom
from functions import get_waypoint, get_route, get_route_point_order
from objects import *

# Load waypoint data from Excel
def load_waypoints(file_path):
    wb = load_workbook(file_path)
    waypoint_table = wb['waypoint-table']
    waypoints = []

    for row in waypoint_table.iter_rows(min_row=2):  # Skip header
        name, identifier, waypoint_type, country_code, lat, lon, comment = [cell.value for cell in row]
        if identifier != "identifier":  # Avoid header repetition
            waypoints.append(Waypoint(name=name, identifier=identifier, waypoint_type=waypoint_type,
                                      country_code=country_code, lat=lat, lon=lon))
    return waypoints


# Load route data from Excel
def load_routes(file_path):
    wb2 = load_workbook(file_path)
    info_table = wb2['info']
    route_points_table = wb2['routes']
    
    routes = []
    for row in info_table.iter_rows(min_row=2):  # Skip header
        flight_plan_index, route_name, route_description = [cell.value for cell in row]
        if flight_plan_index != "flight-plan-index":
            routes.append(Route(route_name=route_name, flight_plan_index=flight_plan_index,
                                route_description=route_description))
    
    # Map route points to routes
    for row in route_points_table.iter_rows(min_row=2):  # Skip header
        flight_plan_index, identifier, order = [cell.value for cell in row]
        waypoint = get_waypoint(identifier=identifier, all_waypoint=all_waypoints)
        if waypoint:
            route = get_route(flight_plan_index=flight_plan_index, all_routes=routes)
            if route:
                route_point = RoutePoint(waypoint=waypoint, order=order)
                route.route_points.append(route_point)
            else:
                print(f"Route {flight_plan_index} doesn't exist!")
                quit()
        else:
            print(f"Waypoint {identifier} doesn't exist!")
            quit()

    # Sort route points by order for each route
    for route in routes:
        route.route_points.sort(key=get_route_point_order)

    return routes


# Create Garmin flight plan XML from route objects
def create_flight_plan_xml(routes, output_dir):
    for fpl in routes:
        root = minidom.Document()

        flight_plan = root.createElement('flight-plan')
        flight_plan.setAttribute('xmlns', 'http://www8.garmin.com/xmlschemas/FlightPlan/v1')
        root.appendChild(flight_plan)

        # Created date (assuming fpl has a 'created' field)
        created = root.createElement('created')
        created.appendChild(root.createTextNode(fpl.created))
        flight_plan.appendChild(created)

        # Waypoint table
        waypoint_table = root.createElement('waypoint-table')
        flight_plan.appendChild(waypoint_table)

        for route_point in fpl.route_points:
            waypoint_obj = route_point.waypoint
            waypoint = root.createElement('waypoint')
            waypoint_table.appendChild(waypoint)

            for tag, value in [('identifier', waypoint_obj.identifier), 
                               ('type', waypoint_obj.type),
                               ('country-code', waypoint_obj.country_code), 
                               ('lat', str(waypoint_obj.lat)), 
                               ('lon', str(waypoint_obj.lon)), 
                               ('comment', waypoint_obj.comment)]:
                elem = root.createElement(tag)
                elem.appendChild(root.createTextNode(value))
                waypoint.appendChild(elem)

            if waypoint_obj.elevation:
                elevation = root.createElement('elevation')
                elevation.appendChild(root.createTextNode(waypoint_obj.elevation))
                waypoint.appendChild(elevation)

        # Route
        route_elem = root.createElement('route')
        flight_plan.appendChild(route_elem)

        for tag, value in [('route-name', fpl.route_name),
                           ('route-description', fpl.route_description),
                           ('flight-plan-index', str(fpl.flight_plan_index))]:
            elem = root.createElement(tag)
            elem.appendChild(root.createTextNode(value))
            route_elem.appendChild(elem)

        # Route points
        for route_point in fpl.route_points:
            route_point_elem = root.createElement('route-point')
            route_elem.appendChild(route_point_elem)

            for tag, value in [('waypoint-identifier', route_point.waypoint.identifier), 
                               ('waypoint-type', route_point.waypoint.type), 
                               ('waypoint-country-code', route_point.waypoint.country_code)]:
                elem = root.createElement(tag)
                elem.appendChild(root.createTextNode(value))
                route_point_elem.appendChild(elem)

        # Save XML to file
        save_path = f"{output_dir}/{fpl.flight_plan_index}-{fpl.route_name}.fpl"
        with open(save_path, "w") as f:
            f.write(root.toprettyxml(indent="\t"))

    print("Flight plan generation complete!")


# Main function to load data and generate XMLs
if __name__ == "__main__":
    all_waypoints = load_waypoints('static/waypoints.xlsx')
    all_routes = load_routes('input/flight_plans.xlsx')
    create_flight_plan_xml(all_routes, 'output')
